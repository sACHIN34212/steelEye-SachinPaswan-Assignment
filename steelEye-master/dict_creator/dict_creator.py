import sys
import json
import requests
import time
import xlrd
import boto3
from openpyxl import Workbook, load_workbook
import logging


def lambda_handler(event, context):
    # get url from config.json file
    with open('config.json', 'r') as config_file:
        data = json.load(config_file)

    xlsx_url = data['url']
    split_url = xlsx_url.split('/')
    file_name = split_url[len(split_url)-1]
    response = requests.get(xlsx_url) 

    # Stops executing the function if file download url returns anything other than 200 (status_code)
    if not response.ok:
        sys.exit('Service Unavailable')

    # checks for 'xls' file format
    if file_name[-3:] == 'xls':      
        # Converting xls to xlsx format
        xlsBook = xlrd.open_workbook(file_contents=response.content)
        workbook = Workbook()

        for i in range(0, xlsBook.nsheets):
            xlsSheet = xlsBook.sheet_by_index(i)
            sheet = workbook.active if i == 0 else workbook.create_sheet()
            sheet.title = xlsSheet.name
        
            for row in range(0, xlsSheet.nrows):
                for col in range(0, xlsSheet.ncols):
                    sheet.cell(row=row + 1, column=col + 1).value = xlsSheet.cell_value(row, col) if xlsSheet.cell_value(row, col) != '' else None

        # Saves data file in xlsx format 
        workbook.save(file_name[:-3]+'xlsx')

    #Checks for xlsx file format
    elif file_name[-4:] == 'xlsx':
        with open(file_name, 'w') as f: # Saving and loading the data file
            f.write(response.content)
        workbook = load_workbook(file_name)

    # Stops executing if the file is not in xls or xlsx format  
    else:
        sys.exit('Unsupported file format')

    try:
        sheet = workbook['MICs List by CC'] # loads 'MICs List by CC' sheet
    except KeyError as error:
        sys.exit('Specified sheet is missing') # Stops executing if 'MICs List by CC' sheet is missing

    data_list = []

    # list of all the column names
    key_values = [sheet.cell(row=1, column=i).value for i in range(1, sheet.max_column+1)]  

    # Generating dictionary of all the rows with column name as keys
    for i in range(2, sheet.max_row+1):
        row_dict = {key_values[x-1]: sheet.cell(row=i, column=x).value for x in range(1, sheet.max_column+1)}
        data_list.append(row_dict)

    # Saving json data on s3
    s3 = boto3.resource('s3')
    time_stamp_string = time.strftime("%Y%m%d-%H%M%S")  # To generate unique filename
    s3.Bucket('steeleye-sachin').put_object(Key='data_' + time_stamp_string +'.json', Body=json.dumps(data_list, indent=4))
    logging.info('done')
    logging.debug('https://s3-ap-southeast-1.amazonaws.com/steeleye-sachin/' + 'data_' + time_stamp_string +'.json')
    # url to the json file in s3
    return ('https://s3-ap-southeast-1.amazonaws.com/steeleye-sachin/' + 'data_' + time_stamp_string +'.json')
